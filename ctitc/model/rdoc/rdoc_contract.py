#  -*- coding: utf-8 -*-
# Project: BrandAnalyze
# Author:  yyg
# Created on 2019-08-04
#########################################################
from docx import Document
from docxtpl import DocxTemplate
import os
from openpyxl import load_workbook
import xlrd


import numpy as np
from com.ctitc.bigdata.db.mysqldb import MysqlDB
from com.ctitc.bigdata.common.log.mylog import MyLog
import pandas as pd

from com.ctitc.bigdata.model.rdoc.rdoc_base import RDOCBase

#########################################################
# 商务合同处理
#########################################################
class RDOCContract(RDOCBase):
    # EXCEL_PATH = '/home/rasmode/tensorflow/data/hbzy'
    EXCEL_PATH = 'E:\\project\\spark\\tmp'
    ##########################################
    # 初始化
    ##########################################
    def __init__(self, start_row=6, row_num=26, start_col=1, col_num=20):
        super().__init__(logkey="rdoc")
        self.start_row = start_row
        self.row_num = row_num
        self.start_col = start_col
        self.col_num = col_num
    pass
    ##########################################
    # 主处理函数
    ##########################################
    def main(self, rq=''):
        # 1. 从EXCEL读取信息
        file_name = "2020年上半年终端市场信息采集跟踪与维护项目供应商金额.xlsx"
        data_df = self.getExcelInfo(file_name)
        # 2.处理word模板
        file_name = "2020年上半年终端市场信息采集跟踪与维护项目合同模板.docx"
        self.processWord(file_name, data_df)
        pass
    pass
    ##########################################
    # 处理word模板
    ##########################################
    def processWord(self, file_name="", info_df=None):
        file_path = os.path.join(self.EXCEL_PATH,file_name)
        if(not os.path.exists(file_path)):
            raise Exception("file:" + file_path + " doesnot exist.")
        pass
        pre_filename = 'JZ-F-2019009-'
        for row in info_df.index:
            self.openDocTemplate(file_path)
            id = info_df.loc[row, 'id']
            # 市场
            market = info_df.loc[row, 'market']
            market = market.replace('、','')
            # 供应商
            supplier = info_df.loc[row, 'supplier']
            # 核心终端建设及维护(元/户)
            hxzd_dj = info_df.loc[row, 'hxzd_dj']
            hxzd_dj = format(float(hxzd_dj), '0,.2f')
            # 核心终端建设及维护(户数)
            hxzd_hs = info_df.loc[row, 'hxzd_hs']
            hxzd_hs = format(float(hxzd_hs), '0,.0f')
            # 发票税率
            fpsl = info_df.loc[row, 'fpsl']
            # 总额
            zje = info_df.loc[row, 'zje']
            zje = format(float(zje), '0,.2f')
            contract_name = pre_filename + id + "(" + market + ").docx"
            contract_path = os.path.join(self.EXCEL_PATH,contract_name)
            context = {'market':market,
                       'supplier':supplier,
                       'hxzd_dj':hxzd_dj,
                       'hxzd_hs':hxzd_hs,
                       'fpsl':fpsl,
                       'zje':zje}
            self.tpl.render(context)
            self.saveTemplate(contract_path)
        pass
    pass

    ##########################################
    # 从EXCEL读取信息, 返回DataFrame
    ##########################################
    def getExcelInfo(self, file_name):
        # 1.拷贝模板并命名
        # 1.1 判断文件是否存在
        file_path = os.path.join(self.EXCEL_PATH,file_name)
        if(not os.path.exists(file_path)):
            raise Exception("file:" + file_path + " doesnot exist.")
        pass
        wb = load_workbook(file_path,data_only=True)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        print(sheetnames)
        # 获取总行数,包括标题
        nrows = sheet.max_row
        print(" 总行数= %s"%nrows)
        # 获取总列数
        ncols = sheet.max_column
        print(" 总列数= %s"%ncols)
        # ncols = 19
        print(" 总列数= %s"%ncols)
        # 读取每行数据
        arr_code = []
        try:
            for i in range(self.start_row, self.start_row + self.row_num):
                row = []
                for j in range(self.start_col, self.start_col + self.col_num):
                    id = sheet.cell(row=i, column=1).value
                    item = str(sheet.cell(row=i, column=j).value)
                    if item == 'None':
                        item = 'NaN'
                    pass
                    row.append(item)
                    # if isinstance(int(id),int):
                    #     row.append(str(sheet.cell(row=i, column=j).value))
                    # else:
                    #     break
                    # pass
                pass
                arr_code.append(row)
            pass
        except  Exception as ex:
            print(str(ex))
        pass
        data_df = pd.DataFrame(arr_code,
                               columns=['id','market','contract_id','supplier','hxzd_dj','hxzd_hs','fpsl','zje',
                                        'first_je','first_kpsj','first_zt','second_je','second_kpsj','second_zt',
                                        'third_je','third_kpsj','third_zt','fourth_je','fourth_kpsj','fourth_zt'])
        values = {'contract_id': '', 'supplier': '', 'hxzd_dj': '0', 'hxzd_hs': '0', 'fpsl': '0',
                  'zje': '0', 'first_je': '0', 'first_kpsj': '', 'first_zt': '','second_je': '0', 'second_kpsj': '',
                  'second_zt': '', 'third_je': '0', 'third_kpsj': '', 'third_zt': '',
                  'fourth_je': '0', 'fourth_kpsj': '', 'fourth_zt': ''}

        data_df.fillna(value=values, inplace=True)
        print(data_df)
        return data_df
    pass
pass
if __name__ == "__main__":
    model = RDOCContract()
    model.main()
pass