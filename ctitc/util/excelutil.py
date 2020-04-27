#!/usr/bin/env python
#  -*- coding: utf-8 -*-
# Project: Brand Analysis
# Author:  yyg
# Created on 2017-10-12
##########################################
import os

from openpyxl import load_workbook
import pandas as pd


##########################################
#
# Excel辅助类
#
##########################################
class ExcelUtil():

    ##########################################
    # 初始化
    ##########################################
    def __init__(self):
        pass
    pass

    ##########################################
    # 根据模板格式,生成EXCEL文件,只支持一个sheet
    ##########################################
    @classmethod
    def save_by_one_template(cls, templatename, filename, data, row_start=0, col_start=0):
        # try:
        # 1.拷贝模板并命名
        # 1.1 判断文件是否存在
        if(not os.path.exists(templatename)):
            raise Exception("file:" + templatename + " doesnot exist.")
        pass
        # 打开该excel,并保存原有的格式
        wb = load_workbook(templatename)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]

        for i,row in enumerate(data):
            for j,col in enumerate(row):
                sheet.cell(row=i + row_start,column=j + col_start).value = col
            pass
        pass
        wb.save(filename)
        # except Exception as ex:
        #     print(" 根据模板格式,生成EXCEL文件错误. " + str(ex))
        # pass
    pass
    ##########################################
    # 根据模板格式,生成EXCEL文件,支持多个sheet
    # args为list,存放dict,一个sheet的内容为一个dict,用法如下:
    # args = []
    # dict1 = {"sheet":'Sheet1', "data":list(rst), "row_start":2, "col_start":1}
    # args.append(dict1)
    # dict2 = {"sheet":'Sheet2', "data":list(rst), "row_start":2, "col_start":1}
    # args.append(dict2)
    # ExcelUtil.save_by_multi_template(self.EXCEL_PATH + "khfb_tmplate.xlsx", self.EXCEL_PATH + "khfb1.xlsx", args)
    ##########################################
    @classmethod
    def save_by_multi_template(cls, templatename, filename, args):
        if(not os.path.exists(templatename)):
            raise Exception("file:" + templatename + " doesnot exist.")
        pass
        # 打开该excel,并保存原有的格式
        wb = load_workbook(templatename)
        for param in args:
            sheet = param['sheet']
            data = param['data']
            row_start = param['row_start']
            col_start = param['col_start']
            sheet = wb[sheet]
            for i,row in enumerate(data):
                for j,col in enumerate(row):
                    sheet.cell(row=i + row_start,column=j + col_start).value = col
                pass
            pass
        pass
        wb.save(filename)
    pass
    ##########################################
    # 根据模板格式,生成EXCEL文件,支持多个sheet
    # args为list,存放dict,一个sheet的内容为一个dict,用法如下:
    # args = []
    # dict1 = {"sheet":'Sheet1', "data":list(rst), "row_start":2, "col_start":1}
    # args.append(dict1)
    # dict2 = {"sheet":'Sheet2', "data":list(rst), "row_start":2, "col_start":1}
    # args.append(dict2)
    # ExcelUtil.save_by_multi_template(self.EXCEL_PATH + "khfb_tmplate.xlsx", self.EXCEL_PATH + "khfb1.xlsx", args)
    ##########################################
    @classmethod
    def save_by_copy_templates(cls, templatename, filename, args):
        if(not os.path.exists(templatename)):
            raise Exception("file:" + templatename + " doesnot exist.")
        pass
        # 打开该excel,并保存原有的格式
        wb = load_workbook(templatename)
        sheet_nums = len(args)
        sheetnames = wb.sheetnames
        tmplate_sheet_name = sheetnames[0]
        for index, param in enumerate(args):
            sheet_name = param['sheet']
            data = param['data']
            row_start = param['row_start']
            col_start = param['col_start']
            # copy模板
            tmp_sheet = wb[tmplate_sheet_name]
            # work_sheet = wb[sheetnames[index]]
            next_sheet = wb.copy_worksheet(tmp_sheet)
            next_sheet.title = sheet_name

            for i,row in enumerate(data):
                for j,col in enumerate(row):
                    try:
                        next_sheet.cell(row=i + row_start,column=j + col_start).value = col
                    except(Exception) as ex:
                        print(ex)
                    pass
                pass
            pass
        pass
        tmp_sheet = wb[tmplate_sheet_name]
        wb.remove(tmp_sheet)
        wb.save(filename)
    pass

    ##########################################
    # 保存excel文件
    ##########################################
    @classmethod
    def fast_save(cls, filename, rst, header=True):
        df = pd.DataFrame.from_records(rst)
        # 写EXCEL文件
        df.to_excel(filename, index=False, header=header)
    pass

pass
##########################################
if __name__ == "__main__":
    excel = ExcelUtil()
    datas = [1,2,3]
    for n in datas:
        print(n)
    pass


pass
################################
