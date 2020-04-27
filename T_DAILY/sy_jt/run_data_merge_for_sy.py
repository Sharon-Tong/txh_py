
import docx
from openpyxl import Workbook
from GDZY_REPORT.LOG.MyLog import logger
import os
import pandas as pd
import xlrd
import re
# import tools

dir=r"C:\Users\Tangt\Desktop\上烟集团工作\201812-201906第三方走访月报\2019年上半年重点终端项目2018年12月相关资料\201812第三方月报\12月 上海烟草河南省第三方终端走访月报告.docx"

# f=docx.Document(dir+file)
# for para in f.paragraphs:
#     text = para.text
#     print(text)


def get_docx(file_name):
    d = docx.opendocx(file_name)
    doc = docx.getdocumenttext(d)
    return doc


#写数据
def make_xls(doc_dir,file_name):
    wb = Workbook()
    ws = wb.active

    doc_dt=get_docx(doc_dir)
    n=len(doc_dt)
    logger.get_logger().info("导入的数据有:"+str(n)+"行")
    #按照条件匹配关键字，并写入
    logger.get_logger().info("开始模糊匹配...")
    for i in range(n):
        if doc_dt[i].find('宏观总体分析')>=0 or doc_dt[i].find('全省综合分析')>=0 or doc_dt[i].find('整体宏观描述')>=0:
            hgzt=i
        if doc_dt[i].find('零售户心态')>=0 or doc_dt[i].find('零售户评价')>=0 :
            lshxt=i
        if doc_dt[i].find('货源分配')>=0:
            hyfp=i
    pass

    print(hgzt,lshxt,hyfp)
    # ws.cell(row=1, column= 1).value = '宏观总体分析'
    # ws.cell(row=1, column= 2).value = '零售户心态'

    ws.cell(row=1, column= 1).value = str(doc_dt[hgzt:lshxt])
    ws.cell(row=1, column= 2).value = str(doc_dt[lshxt:hyfp])



    logger.get_logger().info("写入结束，保存结果")
    wb.save(file_name)

pass


key_order = ['宏观总体分析', '零售户心态']

###将各个doc的文件处理成excel
def get_doc(filePath,filePath_to_xlsx):
    for file_name in os.listdir(filePath):
        logger.get_logger().info("正准备处理文件"+file_name)
        xlsx            =   str.split(file_name,'.docx')[0]+'.xlsx'
        xlsx_path_name  =   os.path.join(filePath_to_xlsx,xlsx)
        doc_dir         =   os.path.join(filePath,file_name)

        logger.get_logger().info("word的路径是：" + doc_dir)
        #make_xls(doc_dir,xlsx_path_name)


###将结果合并
def merge(filePath,filePath_to_xlsx,xlsx_file_anme):
    get_doc(filePath,filePath_to_xlsx)
    #exlcel_name=r'E:\上烟集团工作\excel'
    wb = Workbook()
    ws = wb.active
    i=0
    for filename in os.listdir(filePath_to_xlsx):
        i += 1
        logger.get_logger().info("正准备读取excel文件"+filename)
        #df = pd.read_excel(os.path.join(filePath_to_xlsx, filename))
        data = xlrd.open_workbook(os.path.join(filePath_to_xlsx, filename))
        table = data.sheet_by_name('Sheet')
        # print(table)
        # print( table.cell(0,0).value)
        # print( table.cell(1,2).value)
        ws.cell(row=i, column=1).value = str(filename)
        ws.cell(row=i, column=2).value = table.cell(0,0).value
        ws.cell(row=i, column=3).value = table.cell(0,1).value

        wb.save(xlsx_file_anme)
        # df_empty = df_empty.append(df, ignore_index=True)
        #logger.get_logger().info("合并成一个dataframe" )
    #return df_empty

#写出excel
# def all_toexcel(doc_filePath,doc_to_xlsx_file,xlsx_file_anme,key_order):
#     logger.get_logger().info("将文件开始写出,路径："+ xlsx_file_anme)
#     data=merge(doc_filePath,doc_to_xlsx_file)
#     print(data)
#     wb = Workbook()
#     ws = wb.active
#     #表头写入
#     for key_index,each_key in enumerate(key_order):
#         ws.cell(row=1, column=key_index + 1).value = each_key
#     n,m=data.shape
#     for j in range(m):
#         ws.cell(row=i+2,column=j+1).value =data.iloc[i,j]
#     wb.save(xlsx_file_anme)

###201912
filePath = r'D:\gdzy_work\上海工作-核心终端市场信息采集\test\2020年2月市场走访团队月报'
filePath_to_xlsx = r'D:\gdzy_work\上海工作-核心终端市场信息采集\excel202002'
all_to_xlsx=r'D:\gdzy_work\上海工作-核心终端市场信息采集202002.xlsx'

merge(filePath,filePath_to_xlsx,all_to_xlsx)
