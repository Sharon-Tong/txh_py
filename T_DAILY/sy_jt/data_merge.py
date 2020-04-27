
import docx
from openpyxl import Workbook
from GDZY_REPORT.LOG.MyLog import logger
import os
import pandas as pd

import re
# import tools

dir=r"C:\Users\Tangt\Desktop\上烟集团工作\201812-201906第三方走访月报\2019年上半年重点终端项目2018年12月相关资料\201812第三方月报\12月 上海烟草河南省第三方终端走访月报告.docx"

# f=docx.Document(dir+file)
# for para in f.paragraphs:
#     text = para.text
#     print(text)


def get_docx(file_name):
    d = docx.opendocx(file_name)
    doc = docx.getdocumenttext(d)
    return doc


#写数据
def make_xls(doc_dir,key_order,file_name):
    wb = Workbook()
    ws = wb.active
    #表头写入
    for key_index,each_key in enumerate(key_order):
        ws.cell(row=1, column=key_index + 1).value = each_key
    #获取docx

    doc_dt=get_docx(doc_dir)
    n=len(doc_dt)
    logger.get_logger().info("导入的数据有:"+str(n)+"行")
    #按照条件匹配关键字，并写入
    logger.get_logger().info("开始模糊匹配...")
    i=2
    for key, value in enumerate(key_order):
        for row_value_index in range(n):
            if doc_dt[row_value_index].find(value) >= 0:
                if value in (['报告月份','报告地区','报告公司','微信粉丝','宏观总体分析','零售户心态','商业货源分配']):
                    if doc_dt[row_value_index].find('：')>0:
                        ws.cell(row=i, column=key+1).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
                    if doc_dt[row_value_index].find(':') > 0:
                        ws.cell(row=i, column=key + 1).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]

                if value in (['问题反馈', '其他工业动向']):
                    ws.cell(row=i, column=key+1).value = str(doc_dt[row_value_index + 1:n])
                if value in ['报告内容']:
                    ws.cell(row=i, column=key+1).value = doc_dt[row_value_index + 1]


    j = 1
    for key, value in enumerate(key_order):

        for row_value_index in range(n):
            if value in ['均线图']:
                if doc_dt[row_value_index].find(value) >= 0:
                    if not (str.isdigit(doc_dt[row_value_index]) or str.isspace(doc_dt[row_value_index])):
                        j=j+1
                        ws.cell(row=j, column=key+1).value = doc_dt[row_value_index - 1]
                        #print(doc_dt[row_value_index - 1])



    for key, value in enumerate(key_order):
        for values in (['市场状况小结''价格:', '价格：', '销量：', '销量:', '下月投放建议', '下月营销建议', '竞品状况', '竞品']):
            if value in values:
                k=1
                for row_value_index in range(n):
                    if doc_dt[row_value_index].find(value) >= 0:
                            k=k+1
                            ws.cell(row=k, column=key+1).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]


    logger.get_logger().info("写入结束，保存结果")
    wb.save(file_name)

pass

#key_order = ['报告月份', '报告地区', '报告公司', '报告内容', '微信粉丝', '均线图', '市场状况小结', '价格:', '价格：','销量：','销量:', '下月投放建议',
#             '下月营销建议', '竞品状况', '竞品', '宏观总体分析', '零售户心态', '商业货源分配', '问题反馈'  ,'其他工业动向']

key_order = ['报告月份', '报告地区', '报告公司','宏观总体分析', '零售户心态']

###将各个doc的文件处理成excel
def get_doc(filePath,filePath_to_xlsx):
    for file_name in os.listdir(filePath):
        logger.get_logger().info("正准备处理文件"+file_name)
        xlsx            =   str.split(file_name,'.docx')[0]+'.xlsx'
        xlsx_path_name  =   os.path.join(filePath_to_xlsx,xlsx)
        doc_dir         =   os.path.join(filePath,file_name)
        # if os.path.exists(xlsx_path_name):
        #     pass
        # else:
        #     os.makedirs(xlsx_path_name)
        # logger.get_logger().info("word目录的path是：" + filePath)
        # logger.get_logger().info("word目文件名是：" + file_name)
        logger.get_logger().info("word的路径是：" + doc_dir)
        make_xls(doc_dir, key_order,xlsx_path_name)


###将结果合并
def merge(filePath,filePath_to_xlsx):
    get_doc(filePath,filePath_to_xlsx)
    #exlcel_name=r'E:\上烟集团工作\excel'
    df_empty=pd.DataFrame(columns=key_order)
    for filename in os.listdir(filePath_to_xlsx):
        logger.get_logger().info("正准备读取excel文件"+filename)
        df = pd.read_excel(os.path.join(filePath_to_xlsx, filename))
        df_empty = df_empty.append(df, ignore_index=True)
        logger.get_logger().info("合并成一个dataframe" )
    return df_empty

#写出excel
def all_toexcel(doc_filePath,doc_to_xlsx_file,xlsx_file_anme,key_order):
    logger.get_logger().info("将文件开始写出,路径："+ xlsx_file_anme)
    data=merge(doc_filePath,doc_to_xlsx_file)
    wb = Workbook()
    ws = wb.active
    #表头写入
    for key_index,each_key in enumerate(key_order):
        ws.cell(row=1, column=key_index + 1).value = each_key
    n,m=data.shape
    for i in range(n):
        for j in range(m):
            ws.cell(row=i+2,column=j+1).value =data.iloc[i,j]
    wb.save(xlsx_file_anme)

###201912
filePath = r'D:\gdzy_work\上海工作-核心终端市场信息采集\test\2019年12月市场走访团队月报'
filePath_to_xlsx = r'D:\gdzy_work\上海工作-核心终端市场信息采集\excel'
all_to_xlsx=r'D:\gdzy_work\上海工作-核心终端市场信息采集201912.xlsx'

all_toexcel(filePath,filePath_to_xlsx,all_to_xlsx,key_order)
#
# #201901
# filePath1=r"E:\上烟集团工作\201812-201906第三方走访月报\2019年上半年重点终端项目2019年1月相关资料\第三方月报201901"
# filePath_to_xlsx1 = r'E:\上烟集团工作\excel\201901'
# all_to_xlsx1=r'E:\上烟集团工作\合并\201901.xlsx'
# all_toexcel(filePath1,filePath_to_xlsx1,all_to_xlsx1,key_order)
# #201903
# filePath3=r"E:\上烟集团工作\201812-201906第三方走访月报\2019年上半年重点终端项目2019年3月相关资料\市场走访团队月报201903"
# filePath_to_xlsx3 = r'E:\上烟集团工作\excel\201903'
# all_to_xlsx3=r'E:\上烟集团工作\合并\201903.xlsx'

#all_toexcel(filePath3,filePath_to_xlsx3,all_to_xlsx3,key_order)

#doc_test = make_xls(dir, key_order, 'E:/test_excel.xlsx')

logger.get_logger().info("程序结束")

    #
    #     if doc_dt[row_value_index].find('报告月份')>=0:
    #         i+=1
    #         ws.cell(row=i,column=1).value=doc_dt[row_value_index][(doc_dt[row_value_index].find('：')+1):]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('报告地区')>=0:
    #         i+=1
    #         ws.cell(row=i,column=2).value=doc_dt[row_value_index][(doc_dt[row_value_index].find('：')+1):]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('报告公司')>=0:
    #         i+=1
    #         ws.cell(row=i,column=3).value=doc_dt[row_value_index][(doc_dt[row_value_index].find('：')+1):]
    # for row_value_index in range(n):
    #     i=1
    #     if doc_dt[row_value_index].find('报告内容')>=0:
    #         i+=1
    #         ws.cell(row=i,column=4).value=doc_dt[row_value_index+1]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('微信粉丝')>=0:
    #         i+=1
    #         ws.cell(row=i, column=5).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('均线图')>=0:
    #         if not(str.isdigit(doc_dt[row_value_index]) or str.isspace(doc_dt[row_value_index])):
    #             i+=1
    #             ws.cell(row=i, column=6).value = doc_dt[row_value_index-1]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('市场状况小结')>=0:
    #         i+=1
    #         ws.cell(row=i, column=7).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('价格')>=0:
    #         i+=1
    #         ws.cell(row=i, column=8).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('销量')>=0:
    #         i+=1
    #         ws.cell(row=i, column=9).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i =1
    #     if doc_dt[row_value_index].find('下月投放建议')>=0:
    #         i+=1
    #         ws.cell(row=i, column=10).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i = 1
    #     if doc_dt[row_value_index].find('下月营销建议')>=0:
    #         i+=1
    #         ws.cell(row=i, column=11).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i = 1
    #     if doc_dt[row_value_index].find('竞品状况')>=0:
    #         i+=1
    #         ws.cell(row=i, column=12).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i = 1
    #     if doc_dt[row_value_index].find('竞品')>=0:
    #         i+=1
    #         ws.cell(row=i, column=13).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i = 1
    #     if doc_dt[row_value_index].find('宏观总体分析')>=0:
    #         i+=1
    #         ws.cell(row=i, column=14).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i = 1
    #     if doc_dt[row_value_index].find('零售户心态')>=0:
    #         i+=1
    #         ws.cell(row=i, column=15).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i = 1
    #     if doc_dt[row_value_index].find('商业货源分配')>=0:
    #         i+=1
    #         ws.cell(row=i, column=16).value = doc_dt[row_value_index][(doc_dt[row_value_index].find('：') + 1):]
    # for row_value_index in range(n):
    #     i = 1
    #     if doc_dt[row_value_index].find('问题反馈')>=0:
    #         ws.cell(row=i, column=17).value = str(doc_dt[row_value_index+1:n])
    #     # if doc_dt[row_value_index].find('其他工业动向')>=0:
    #     #     ws.cell(row=i, column=18).value = doc_dt[row_value_index+1:n]
    # pass

#     logger.get_logger().info("写入结束，保存结果")
#     wb.save(file_name)
# pass
#
# key_order=['月份','地域','公司','内容','微信粉丝','规格','市场状况小结','价格','销量','下月投放建议',
#            '下月营销建议','竞品状况','竞品','宏观总体分析','零售户心态','商业货源分配','问题反馈'#,'其他工业动向'
#            ]
#
# doc_test=make_xls(dir,key_order,'E:/test_excel.xlsx')
# logger.get_logger().info("程序结束")
# # doc_test=get_docx(dir)
# # for para in doc_test:
# #     if para.find('')
# #     print(para)