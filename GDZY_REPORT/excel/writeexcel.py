# -*- coding: utf-8 -*-

'''
 Time : 2020/3/30 8:22
 Author : SharonTong
'''
from openpyxl import Workbook
from openpyxl.styles import Font, colors, PatternFill
from openpyxl.styles import Border, Side,Alignment
from GDZY_REPORT.LOG.MyLog import logger

#######################################
    # sql查询语句
    # 操作excel
    #######################################
class To_excel():

    def __init__(self,GdzyParam):
        self.configParam=GdzyParam
        self.logger=logger.get_logger()

    def save_data_2_excel(self,data):

        #读取数据
        #data = self.objData.select_get_original_data( self.configParam['startdate'], self.configParam['enddate'] , self.configParam['last_month'], self.dataBase , self.configParam['key_order'] )
        #关闭数据库
        #self.objData.endprocess_databse( self.dataBase );

        #Logger().logger.info( u'需要处理的数据量：' + str(len(data)) )

        #操作excel
        self.logger.info('开始写入excel...')
        filename   = self.configParam.filename
        wb         = Workbook()
        ws         = wb.active
        ws.title   = self.configParam.title

        # 调整颜色
        r_single_fill = PatternFill(fill_type='solid', fgColor=colors.WHITE)
        r_double_fill = PatternFill(fill_type='solid', fgColor='ADEAEA')

        # 向ws写数据
        i = 1

        #表头写入
        #单元格边框
        border_set = Border(right=Side(style='thin', color=colors.BLACK),
                            bottom=Side(style='thin', color=colors.BLACK))
        key_order = self.configParam.key_order

        for index_key, each_keyvalue in enumerate(self.configParam.key_order):

            ws.cell(row= 1 , column=index_key+1).value = each_keyvalue

            ws.cell(row=1, column=index_key + 1).border = border_set
            ws.cell(row=1, column=index_key + 1).alignment = Alignment(horizontal='center', vertical='center')

        pass
        ws.freeze_panes = 'A2'  # 冻结第一行
        #将单位写入
        ws.cell(row=1, column=len(key_order) + 1).value = u"单位：箱、万元"

        #每行数据写入
        #单元格边框
        border_side = Border(right=Side(style='thin', color=colors.BLACK))

        for each_data in data:

            i = i + 1

            # 双数行颜色变为浅蓝色，单数行为白色，单元格边框设置为右画线
            if i % 2 != 0:

                for index, each_key in enumerate(each_data):

                    colnum = index + 1
                    ws.cell(row=i, column=colnum).border = border_side

                    ws.cell(row=i, column=colnum).fill = r_single_fill

                    if str(each_data[each_key]).replace("-", "").replace(".", "").isdigit():
                        ws.cell(row=i, column=colnum).value = float(each_data[each_key])

                    else:
                        ws.cell(row=i, column=colnum).value = each_data[each_key]
                    pass
                pass
            else:
                for index, each_key in enumerate(each_data):
                    colnum = index + 1
                    ws.cell(row=i, column=colnum).border = border_side

                    ws.cell(row=i, column=colnum).fill = r_double_fill

                    if str(each_data[each_key]).replace("-", "").replace(".", "").isdigit():
                        ws.cell(row=i, column=colnum).value = float(each_data[each_key])

                    else:
                        ws.cell(row=i, column=colnum).value = each_data[each_key]

                    pass
                pass

            pass

        pass

        #字体
        font = Font(name='微软雅黑', size=10)
        #align = Alignment(horizontal=’center’, vertical =’center’)

        #最后一行增加下划线
        row_max = ws.max_row
        border_bottom = Border(right=Side(style='thin', color=colors.BLACK),
                               bottom=Side(style='thin', color=colors.BLACK))

        for index, each_key in enumerate( self.configParam.key_order ):

            ws.cell(row=row_max, column=index+1).border = border_bottom

        pass


        #调整行高
        ws.row_dimensions[1].height = 40.0
        #保存工作簿
        wb.save(filename)
        self.logger.info('写入excel结束...')

# if __name__== "__main__":
#     a=MyDB2(db88)
#     a.connect()
#     data=a.select("select * from DBREAD.DIM_PRICE_PUBLIC fetch first 2 rows only")
#     a.close()
#     b=To_excel(DailyParam)
#     b.save_data_2_excel(data)