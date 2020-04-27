
# -*- coding: utf-8 -*-

'''
 Time : 2020/3/30 8:22
 Author : SharonTong
'''
from openpyxl import Workbook
from openpyxl.styles import Font, colors, PatternFill
from openpyxl.styles import Border, Side
from GDZY_REPORT.T_gd_z_monthly_sale.a_monthly_param import DailyParam
from GDZY_REPORT.LOG.MyLog import logger
from GDZY_REPORT.Sql_param.sqlconfig import db88
from GDZY_REPORT.SQL_LINK.DB2_link import MyDB2
#######################################
    # sql查询语句
    # 操作excel
    #######################################
class To_excel():

    def __init__(self,GdzyParam):
        self.configParam=GdzyParam
        self.logger=logger.get_logger()

    def save_data_2_excel(self,data1,data2):

        #读取数据
        #data = self.objData.select_get_original_data( self.configParam['startdate'], self.configParam['enddate'] , self.configParam['last_month'], self.dataBase , self.configParam['key_order'] )
        #关闭数据库
        #self.objData.endprocess_databse( self.dataBase );

        #Logger().logger.info( u'需要处理的数据量：' + str(len(data)) )

        #操作excel,创建2个工作簿
        filename   = self.configParam.filename
        wb         = Workbook()
        ws         = wb.active
        ws.title   = self.configParam.title_prov
        ws1        = wb.create_sheet(title=self.configParam.title_dept)
        self.logger.info("开始将省五率写入exccel...")
        # 调整颜色
        r_single_fill = PatternFill(fill_type='solid', fgColor=colors.WHITE)
        #r_double_fill = PatternFill(fill_type='solid', fgColor='ADEAEA')

        # 向ws写数据
        i = 2

        #表头写入
        #单元格边框
        border_set = Border(right=Side(style='thin', color=colors.BLACK),
                            bottom=Side(style='thin', color=colors.BLACK),
                            left=Side(style='thin', color=colors.BLACK),
                            top=Side(style='thin', color=colors.BLACK))


        #合并单元格
        n   =   len(self.configParam.key_order_prov)
        m   =   len(self.configParam.key_order_dept)

        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n)
        ws1.merge_cells(start_row=1, end_row=1, start_column=1, end_column=m)

        #第一行写入
        ws.cell(row=1, column=1).value = self.configParam.key_order_prov_t
        ws.cell(row=1, column=1).border = border_set

        ws1.cell(row=1, column=1).value = self.configParam.key_order_dept_t
        ws1.cell(row=1, column=1).border = border_set

        #第二行写入
        for index_key, each_keyvalue in enumerate(self.configParam.key_order_prov):

            ws.cell(row= 2 , column=index_key+1).value = each_keyvalue

            ws.cell(row=2, column=index_key + 1).border = border_set

        pass


        for each_data in data1:
            i = i + 1

            for index, each_key in enumerate(each_data):

                colnum = index + 1
                ws.cell(row=i, column=colnum).border = border_set

                if str(each_data[each_key]).replace("-", "").replace(".", "").isdigit():
                    ws.cell(row=i, column=colnum).value = float(each_data[each_key])

                else:
                    ws.cell(row=i, column=colnum).value = each_data[each_key]
                pass
            pass
        pass
        self.logger.info('共写入数据：'+str(len(data1)))
        self.logger.info('省五率写入excel表已完成...')

        # 第二行写入
        self.logger.info("开始将地市五率写入exccel...")
        for index_key, each_keyvalue in enumerate(self.configParam.key_order_dept):
            ws1.cell(row=2, column=index_key + 1).value = each_keyvalue

            ws1.cell(row=2, column=index_key + 1).border = border_set

        pass
        j=2
        for each_data in data2:
            j = j + 1

            for index, each_key in enumerate(each_data):

                colnum = index + 1
                ws1.cell(row=j, column=colnum).border = border_set

                if str(each_data[each_key]).replace("-", "").replace(".", "").isdigit():
                    ws1.cell(row=j, column=colnum).value = float(each_data[each_key])

                else:
                    ws1.cell(row=j, column=colnum).value = each_data[each_key]
                pass
            pass
        pass
        self.logger.info('共写入数据：' + str(len(data2)))
        self.logger.info('地市五率写入excel表已完成...')


        #字体
        font = Font(name='微软雅黑', size=10)
        #align = Alignment(horizontal=’center’, vertical =’center’)

        #调整行高
        ws.row_dimensions[1].height = 20.0
        ws1.row_dimensions[1].height = 20.0

        #保存工作簿
        wb.save(self.configParam.filename)

# if __name__== "__main__":
#     a=MyDB2(db88)
#     a.connect()
#     data=a.select("select * from DBREAD.DIM_PRICE_PUBLIC fetch first 2 rows only")
#     b=To_excel(DailyParam)
#     b.save_data_2_excel(data,data)
#     a.close()